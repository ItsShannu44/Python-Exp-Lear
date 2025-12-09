import openpyxl as op

wb=op.load_workbook('VideoSales.xlsx')
print(type(wb))

print(wb)

#Call the Active Worksheet
ws=wb.active
print(type(ws))
print(ws)

#Call the sheet by name
ws=wb['SalesData']
print(ws)


#Count the num of rows and cols is used in worksheet

print(f'Total number of rows is {ws.max_row} and total number of columns is {ws.max_column}.')

#Read data from cell
print(f'The value stored in the cell D5 is : {ws['D5'].value}')

#Read data from multiple cells
values=[ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
print(values)

#Add a new row
n_r=(31,'Cricket Premier league','PC',2025,'Sports','GameStudio',2.60, 1.40, 3.80, 2.20, 10.0)
ws.append(n_r)
wb.save('VideoSales.xlsx')
print(ws.max_row)

#Delete a row
ws.delete_rows(ws.max_row,1)
print(ws.max_row)



#---------------------EXCEL FORMULAS---------------------------------
ws['M1']='Average Sales'
ws['M2']='=AVERAGE(J2:J35)'
wb.save('VideoSales.xlsx')

#Count 'a'
n_r=(31,'Cricket Premier league','PC',' ','Sports','GameStudio',2.60, 1.40, 3.80, 2.20, 10.0)
ws.append(n_r)
wb.save('VideoSales.xlsx')

ws['N1']='Number of rows that have the value'
ws['N2']='=COUNTA(D1:D34)'
wb.save('VideoSales.xlsx')

#Count if

# ws['01']='Num of rows with sports genre'
# ws['02']='=COUNTIF(E2:E34, "Sports")'
# wb.save('VideoSales.xlsx')

#Sumif
ws['M5']='Total Sports sales'
ws['M6']='=SUMIF(E2:E34, "Sports", K2:K34)'
wb.save('VideoSales.xlsx')


#------------------Creation of new worksheet-----------------

wb.create_sheet('New Sheet')
wb.save('VideoSales.xlsx')

print(wb.sheetnames)

#Title of active sheet using ws.title
print(ws.title)

#Rename a sheet
# ws=wb['A New Sheet1']
# print(ws.title)
# ws.title='SalesData'
# wb.save('VideoSales.xlsx')

#Duplicate a worksheet
# wb.copy_worksheet(wb['SalesData'])
# wb.save('VideoSales.xlsx')
# print(wb.sheetnames)

